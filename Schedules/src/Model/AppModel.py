import sqlite3

from PyQt5 import QtSql
from PyQt5.QtCore import pyqtSignal, QObject
from openpyxl import load_workbook, Workbook

from src.Model.DataBasesEnum import DataBasesEnum
from src.Model.Schedule import Schedule
from src.Model.DataBaseWorker import DataBaseWorker
from src.Model.Subject import Subject


class AppModel(QObject):
    tables_updated = pyqtSignal()

    def __init__(self):
        super().__init__()

        self.currentDriver = "QSQLITE"
        self._dataBaseWorker = DataBaseWorker(QtSql.QSqlDatabase.addDatabase(self.currentDriver))

        self.day_row = {
          'понедельник': range(2, 16),
          'вторник': range(16, 30),
          'среда': range(30, 44),
          'четверг': range(44, 58),
          'пятница': range(58, 72),
          'суббота': range(72, 86)
        }

        self._listTables: [str] = []
        self._set_connections()

    def _set_connections(self):
        self._dataBaseWorker.tables_updated.connect(self._on_tables_updated)

    def _on_tables_updated(self):
        self._listTables = self._dataBaseWorker.get_tables()
        self.tables_updated.emit()

    def connect_to_db(self, config: {}) -> bool:
        if config["database_driver"] == DataBasesEnum.Mysql:
            self.currentDriver = "QMYSQL"
        elif config["database_driver"] == DataBasesEnum.Sqlite:
            self.currentDriver = "QSQLITE"

        return self._dataBaseWorker.connect_to_database(config)

    def get_tables(self) -> [str]:
        return self._listTables

    @staticmethod
    def load_workbook(pathToFile: str) -> Workbook:
        try:
            workbook = load_workbook(pathToFile)

            return workbook
        except BaseException as error:
            raise Exception("Error load workbook")

    @staticmethod
    def save_workbook(pathToFile: str, workbook: Workbook):
        try:
            workbook.save(pathToFile)
        except BaseException as error:
            raise Exception('Error writing workbook')

    def load_data_from_workbook(self, workbook: Workbook) -> [Schedule]:
        worksheet = workbook.active
        schedules = []

        for col in worksheet.iter_cols(min_row=1, min_col=4,
                                       max_row=worksheet.max_row,
                                       max_col=worksheet.max_column):
            group = col[0].value.strip()

            for cell in col:
                if (cell.value is None) or (cell.value == '') or (cell.row == 1):
                    continue

                data = cell.value.split(',')

                subject = data[0]
                teacher = data[1]
                lecture = data[2]
                typeSubject = data[3].replace(" ", "")

                subject = Subject(subject, lecture, teacher, typeSubject)

                day = [key for (key, value) in self.day_row.items()
                       if cell.row in value][0]

                index = self.day_row[day].index(cell.row)
                subject_number = (index // 2) + 1
                week = 2 if index % 2 else 1

                schedules.append(Schedule(group, week, subject, day, subject_number))
        return schedules

    def add_schedules_to_database(self, schedules):
        if not self._dataBaseWorker.is_connected():
            raise Exception("Database not connected")

        for schedule in schedules:
            self._dataBaseWorker.add_new_schedule(schedule)

    def export_database_to_sql(self, file_path: str):
        if not self._dataBaseWorker.is_connected():
            raise Exception("Database not connected")

        if self.currentDriver == "QSQLITE":
            connection = sqlite3.connect(self._dataBaseWorker.get_database_name())
            with open(file_path, 'w') as file:
                for line in connection.iterdump():
                    file.write('%s\n' % line)

    def load_database_from_sql(self, file_path):
        if not self._dataBaseWorker.is_connected():
            raise Exception("Database not connected")

    def export_database_as_workbook(self):
        if not self._dataBaseWorker.is_connected():
            raise Exception("Database not connected")

